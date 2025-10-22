import logging
import logging.config
import sqlite3
import os
import asyncio
import shutil
import traceback
import re
import sys
import aiosqlite
import json
import pandas as pd
import io
import aiofiles
from datetime import datetime, timedelta
from contextlib import asynccontextmanager
from typing import Dict, List, Optional, Callable, Any, Awaitable, Union
from aiogram import Bot, Dispatcher, types
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram import F
from aiogram.utils.keyboard import ReplyKeyboardBuilder, InlineKeyboardBuilder
from aiogram.types import ReplyKeyboardRemove
from aiogram.filters import Command
from aiogram.client.default import DefaultBotProperties
from aiogram import BaseMiddleware
from dotenv import load_dotenv

# Загрузка переменных окружения
load_dotenv()

# Настройки
API_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
ADMIN_ID = int(os.getenv('ADMIN_ID', '5024165375'))

# Проверка обязательных переменных
if not API_TOKEN:
    logging.error("Токен бота не найден! Установите переменную TELEGRAM_BOT_TOKEN")
    exit(1)

# ========== УЛУЧШЕНИЕ: РАСШИРЕННОЕ ЛОГИРОВАНИЕ ==========
def setup_logging():
    """Расширенная настройка логирования"""
    logging.config.dictConfig({
        'version': 1,
        'disable_existing_loggers': False,
        'formatters': {
            'detailed': {
                'format': '%(asctime)s - %(name)s - %(levelname)s - %(message)s - [%(filename)s:%(lineno)d]'
            },
            'simple': {
                'format': '%(levelname)s - %(message)s'
            }
        },
        'handlers': {
            'file': {
                'class': 'logging.handlers.RotatingFileHandler',
                'filename': 'bot.log',
                'maxBytes': 10*1024*1024,  # 10MB
                'backupCount': 5,
                'formatter': 'detailed',
                'encoding': 'utf-8'
            },
            'console': {
                'class': 'logging.StreamHandler',
                'formatter': 'simple'
            }
        },
        'loggers': {
            '': {
                'handlers': ['file', 'console'],
                'level': 'INFO'
            }
        }
    })

setup_logging()

# ========== УЛУЧШЕНИЕ: ОНЛАЙН EXCEL С СИНХРОНИЗАЦИЕЙ ==========
class OnlineExcelManager:
    def __init__(self):
        self.active_sessions: Dict[int, Dict] = {}  # user_id -> session_data
        self.sync_interval = 300  # 5 минут
        self.auto_save_task = None
    
    async def start_auto_save(self):
        """Запуск автоматического сохранения"""
        if not self.auto_save_task:
            self.auto_save_task = asyncio.create_task(self._auto_save_loop())
    
    async def _auto_save_loop(self):
        """Цикл автоматического сохранения"""
        while True:
            try:
                await asyncio.sleep(self.sync_interval)
                await self.sync_all_sessions()
            except Exception as e:
                logging.error(f"Ошибка в авто-сохранении: {e}")
    
    async def create_session(self, user_id: int, filters: List[Dict]) -> str:
        """Создание сессии онлайн Excel"""
        try:
            # Создаем DataFrame
            df = self._create_dataframe(filters)
            
            # Сохраняем в файл
            filename = f"online_excel_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = f"temp/{filename}"
            
            # Создаем временную директорию если нужно
            os.makedirs("temp", exist_ok=True)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Фильтры', index=False)
                
                # Настраиваем форматирование
                worksheet = writer.sheets['Фильтры']
                self._apply_excel_formatting(worksheet, df)
            
            # Сохраняем сессию
            self.active_sessions[user_id] = {
                'filepath': filepath,
                'filename': filename,
                'filters': filters,
                'last_update': datetime.now(),
                'is_modified': False
            }
            
            return filepath
            
        except Exception as e:
            logging.error(f"Ошибка создания сессии Excel: {e}")
            raise
    
    def _create_dataframe(self, filters: List[Dict]) -> pd.DataFrame:
        """Создание DataFrame из фильтров"""
        today = datetime.now().date()
        data = []
        
        for f in filters:
            expiry_date = datetime.strptime(str(f['expiry_date']), '%Y-%m-%d').date()
            last_change = datetime.strptime(str(f['last_change']), '%Y-%m-%d').date()
            days_until_expiry = (expiry_date - today).days
            
            icon, status = get_status_icon_and_text(days_until_expiry)
            
            data.append({
                'ID': f['id'],
                'Тип фильтра': f['filter_type'],
                'Местоположение': f['location'],
                'Дата последней замены': last_change.strftime('%d.%m.%Y'),
                'Дата истечения срока': expiry_date.strftime('%d.%m.%Y'),
                'Осталось дней': days_until_expiry,
                'Статус': status,
                'Иконка статуса': icon,
                'Срок службы (дни)': f['lifetime_days'],
                'Дата создания': f['created_at'],
                'Последнее обновление': f.get('updated_at', '')
            })
        
        return pd.DataFrame(data)
    
    def _apply_excel_formatting(self, worksheet, df):
        """Применение форматирования к Excel файлу"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Заголовки
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
        
        # Настройка ширины колонок
        for col_num, column_title in enumerate(df.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for row_num in range(1, worksheet.max_row + 1):
                cell_value = str(worksheet[f"{column_letter}{row_num}"].value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Условное форматирование для статусов
        self._apply_conditional_formatting(worksheet, df)
    
    def _apply_conditional_formatting(self, worksheet, df):
        """Применение условного форматирования"""
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import PatternFill
        
        # Красный для просроченных (колонка F - "Осталось дней")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_rule = FormulaRule(formula=['$F2<=0'], fill=red_fill)
        worksheet.conditional_formatting.add(f"F2:F{len(df)+1}", red_rule)
        
        # Желтый для скоро истекающих
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        yellow_rule = FormulaRule(formula=['AND($F2>0, $F2<=7)'], fill=yellow_fill)
        worksheet.conditional_formatting.add(f"F2:F{len(df)+1}", yellow_rule)
        
        # Оранжевый для предупреждения
        orange_fill = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
        orange_rule = FormulaRule(formula=['AND($F2>7, $F2<=30)'], fill=orange_fill)
        worksheet.conditional_formatting.add(f"F2:F{len(df)+1}", orange_rule)
    
    async def update_session(self, user_id: int, new_filters: List[Dict]) -> bool:
        """Обновление сессии новыми данными"""
        if user_id not in self.active_sessions:
            return False
        
        try:
            session = self.active_sessions[user_id]
            session['filters'] = new_filters
            session['last_update'] = datetime.now()
            session['is_modified'] = True
            
            # Пересоздаем файл
            df = self._create_dataframe(new_filters)
            with pd.ExcelWriter(session['filepath'], engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Фильтры', index=False)
                worksheet = writer.sheets['Фильтры']
                self._apply_excel_formatting(worksheet, df)
            
            return True
        except Exception as e:
            logging.error(f"Ошибка обновления сессии: {e}")
            return False
    
    async def sync_session(self, user_id: int) -> bool:
        """Синхронизация сессии с текущими данными из БД"""
        if user_id not in self.active_sessions:
            return False
        
        try:
            current_filters = await get_user_filters(user_id)
            return await self.update_session(user_id, current_filters)
        except Exception as e:
            logging.error(f"Ошибка синхронизации сессии: {e}")
            return False
    
    async def sync_all_sessions(self):
        """Синхронизация всех активных сессий"""
        for user_id in list(self.active_sessions.keys()):
            try:
                await self.sync_session(user_id)
                logging.info(f"Сессия пользователя {user_id} синхронизирована")
            except Exception as e:
                logging.error(f"Ошибка синхронизации сессии {user_id}: {e}")
    
    async def get_session_file(self, user_id: int) -> Optional[str]:
        """Получение пути к файлу сессии"""
        if user_id in self.active_sessions:
            return self.active_sessions[user_id]['filepath']
        return None
    
    async def close_session(self, user_id: int):
        """Закрытие сессии и удаление файла"""
        if user_id in self.active_sessions:
            try:
                filepath = self.active_sessions[user_id]['filepath']
                if os.path.exists(filepath):
                    os.remove(filepath)
                del self.active_sessions[user_id]
            except Exception as e:
                logging.error(f"Ошибка закрытия сессии: {e}")
    
    async def cleanup_old_sessions(self):
        """Очистка старых сессий (старше 24 часов)"""
        now = datetime.now()
        users_to_remove = []
        
        for user_id, session in self.active_sessions.items():
            if (now - session['last_update']).total_seconds() > 86400:  # 24 часа
                users_to_remove.append(user_id)
        
        for user_id in users_to_remove:
            await self.close_session(user_id)

# Инициализация менеджера Excel
excel_manager = OnlineExcelManager()

# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========
def get_main_keyboard():
    """Клавиатура главного меню"""
    builder = ReplyKeyboardBuilder()
    builder.button(text="📋 Мои фильтры")
    builder.button(text="✨ Добавить фильтр")
    builder.button(text="⚙️ Управление фильтрами")
    builder.button(text="📊 Статистика")
    builder.button(text="📤 Импорт/Экспорт")
    builder.adjust(2)
    return builder.as_markup(resize_keyboard=True)

def get_add_filter_keyboard():
    """Клавиатура для добавления фильтров"""
    builder = ReplyKeyboardBuilder()
    builder.button(text="➕ Один фильтр")
    builder.button(text="📝 Несколько фильтров")
    builder.button(text="🔙 Назад")
    builder.adjust(2)
    return builder.as_markup(resize_keyboard=True)

def get_filter_type_keyboard():
    """Клавиатура для выбора типа фильтра"""
    builder = ReplyKeyboardBuilder()
    builder.button(text="Магистральный SL10")
    builder.button(text="Магистральный SL20")
    builder.button(text="Гейзер")
    builder.button(text="Аквафор")
    builder.button(text="Другой тип фильтра")
    builder.button(text="🔙 Назад")
    builder.adjust(2)
    return builder.as_markup(resize_keyboard=True)

def get_management_keyboard():
    """Клавиатура управления"""
    builder = ReplyKeyboardBuilder()
    builder.button(text="✏️ Редактировать фильтр")
    builder.button(text="🗑️ Удалить фильтр")
    builder.button(text="📊 Онлайн Excel")
    builder.button(text="🔙 Назад")
    builder.adjust(2)
    return builder.as_markup(resize_keyboard=True)

def get_import_export_keyboard():
    """Клавиатура импорта/экспорта"""
    builder = ReplyKeyboardBuilder()
    builder.button(text="📤 Экспорт в Excel")
    builder.button(text="📥 Импорт из Excel")
    builder.button(text="📋 Шаблон Excel")
    builder.button(text="🔙 Назад")
    builder.adjust(2)
    return builder.as_markup(resize_keyboard=True)

def get_back_keyboard():
    """Клавиатура с кнопкой Назад"""
    builder = ReplyKeyboardBuilder()
    builder.button(text="🔙 Назад")
    return builder.as_markup(resize_keyboard=True)

def get_cancel_keyboard():
    """Клавиатура отмены"""
    builder = ReplyKeyboardBuilder()
    builder.button(text="❌ Отмена")
    return builder.as_markup(resize_keyboard=True)

def get_edit_keyboard():
    """Клавиатура для редактирования"""
    builder = ReplyKeyboardBuilder()
    builder.button(text="💧 Тип фильтра")
    builder.button(text="📍 Местоположение")
    builder.button(text="📅 Дата замены")
    builder.button(text="⏱️ Срок службы")
    builder.button(text="🔙 Назад")
    builder.adjust(2)
    return builder.as_markup(resize_keyboard=True)

async def get_online_excel_keyboard():
    """Клавиатура для управления онлайн Excel"""
    builder = InlineKeyboardBuilder()
    builder.button(text="🔄 Синхронизировать", callback_data="sync_excel")
    builder.button(text="📥 Скачать актуальную версию", callback_data="download_excel")
    builder.button(text="📊 Статистика файла", callback_data="excel_stats")
    builder.button(text="❌ Закрыть сессию", callback_data="close_excel")
    builder.adjust(1)
    return builder.as_markup()

def get_status_icon_and_text(days_until_expiry: int):
    """Получение иконки и текста статуса"""
    if days_until_expiry <= 0:
        return "🔴", "ПРОСРОЧЕН"
    elif days_until_expiry <= 7:
        return "🟡", "СКОРО ИСТЕЧЕТ"
    elif days_until_expiry <= 30:
        return "🟠", "ВНИМАНИЕ"
    else:
        return "🟢", "НОРМА"

def format_date_nice(date):
    """Красивое форматирование даты"""
    return date.strftime("%d.%m.%Y")

def create_expiry_infographic(filters):
    """Создание инфографики по срокам"""
    today = datetime.now().date()
    expired = 0
    expiring_soon = 0
    normal = 0
    
    for f in filters:
        expiry_date = datetime.strptime(str(f['expiry_date']), '%Y-%m-%d').date()
        days_until = (expiry_date - today).days
        
        if days_until <= 0:
            expired += 1
        elif days_until <= 7:
            expiring_soon += 1
        else:
            normal += 1
    
    return (
        f"📊 <b>СТАТУС ФИЛЬТРОВ:</b>\n"
        f"🟢 Норма: {normal}\n"
        f"🟡 Скоро истечет: {expiring_soon}\n"
        f"🔴 Просрочено: {expired}"
    )

def is_admin(user_id: int) -> bool:
    """Проверка прав администратора"""
    return user_id == ADMIN_ID

def backup_database() -> bool:
    """Создание резервной копии базы данных"""
    try:
        if os.path.exists('filters.db'):
            backup_name = f'filters_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
            shutil.copy2('filters.db', backup_name)
            return True
    except Exception as e:
        logging.error(f"Ошибка при создании резервной копии: {e}")
    return False

# ========== УЛУЧШЕНИЕ: ВАЛИДАЦИЯ ПОЛЬЗОВАТЕЛЬСКОГО ВВОДА ==========
def validate_filter_type(filter_type: str) -> tuple[bool, str]:
    """Валидация типа фильтра"""
    if not filter_type or len(filter_type.strip()) == 0:
        return False, "Тип фильтра не может быть пустым"
    
    if len(filter_type) > 100:
        return False, "Тип фильтра слишком длинный (макс. 100 символов)"
    
    # Проверка на запрещенные символы
    if re.search(r'[<>{}[\]]', filter_type):
        return False, "Тип фильтра содержит запрещенные символы"
    
    return True, "OK"

def validate_location(location: str) -> tuple[bool, str]:
    """Валидация местоположения"""
    if not location or len(location.strip()) == 0:
        return False, "Местоположение не может быть пустым"
    
    if len(location) > 50:
        return False, "Местоположение слишком длинное (макс. 50 символов)"
    
    if re.search(r'[<>{}[\]]', location):
        return False, "Местоположение содержит запрещенные символы"
    
    return True, "OK"

def validate_lifetime(lifetime: str) -> tuple[bool, str, int]:
    """Валидация срока службы"""
    try:
        days = int(lifetime)
        if days <= 0:
            return False, "Срок службы должен быть положительным числом", 0
        if days > 3650:  # 10 лет
            return False, "Срок службы не может превышать 10 лет", 0
        return True, "OK", days
    except ValueError:
        return False, "Срок службы должен быть числом (дни)", 0

# ========== УЛУЧШЕНИЕ: МОНИТОРИНГ ЗДОРОВЬЯ ==========
class BotHealthMonitor:
    def __init__(self):
        self.start_time = datetime.now()
        self.message_count = 0
        self.error_count = 0
        self.user_actions = {}
    
    def record_message(self, user_id: int):
        """Запись сообщения пользователя"""
        self.message_count += 1
        if user_id not in self.user_actions:
            self.user_actions[user_id] = 0
        self.user_actions[user_id] += 1
    
    def record_error(self):
        """Запись ошибки"""
        self.error_count += 1
    
    async def get_health_status(self):
        """Получение статуса здоровья бота"""
        uptime = datetime.now() - self.start_time
        active_users = len([uid for uid, count in self.user_actions.items() if count > 0])
        
        health_score = (self.message_count - self.error_count) / max(1, self.message_count) * 100
        
        return {
            'uptime': str(uptime),
            'message_count': self.message_count,
            'error_count': self.error_count,
            'active_users': active_users,
            'health_score': health_score
        }

health_monitor = BotHealthMonitor()

# ========== УЛУЧШЕНИЕ: RATE LIMITING ==========
class RateLimiter:
    def __init__(self, max_requests: int = 5, window: int = 30):
        self.max_requests = max_requests
        self.window = window
        self.user_requests = {}
    
    def is_allowed(self, user_id: int) -> bool:
        """Проверка разрешения на обработку запроса"""
        now = datetime.now()
        if user_id not in self.user_requests:
            self.user_requests[user_id] = []
        
        # Удаляем старые запросы
        self.user_requests[user_id] = [
            req_time for req_time in self.user_requests[user_id]
            if (now - req_time).seconds < self.window
        ]
        
        # Проверяем лимит
        if len(self.user_requests[user_id]) >= self.max_requests:
            return False
        
        self.user_requests[user_id].append(now)
        return True

rate_limiter = RateLimiter(max_requests=10, window=30)

# ========== УЛУЧШЕНИЕ: MIDDLEWARE ДЛЯ RATE LIMITING ==========
class RateLimitMiddleware(BaseMiddleware):
    async def __call__(
        self,
        handler: Callable[[types.TelegramObject, Dict[str, Any]], Awaitable[Any]],
        event: types.TelegramObject,
        data: Dict[str, Any]
    ) -> Any:
        # Проверяем, есть ли пользователь
        if hasattr(event, 'from_user') and event.from_user:
            user_id = event.from_user.id
            
            if not rate_limiter.is_allowed(user_id):
                if hasattr(event, 'answer'):
                    await event.answer("⏳ <b>Слишком много запросов!</b>\n\nПожалуйста, подождите 30 секунд.", parse_mode='HTML')
                return
            
            health_monitor.record_message(user_id)
        
        return await handler(event, data)

# Инициализация бота с улучшенными настройками
bot = Bot(
    token=API_TOKEN,
    default=DefaultBotProperties(parse_mode='HTML')
)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)

# Регистрация middleware
dp.update.outer_middleware(RateLimitMiddleware())

# ========== УЛУЧШЕНИЕ: АСИНХРОННАЯ БАЗА ДАННЫХ ==========
@asynccontextmanager
async def get_db_connection():
    """Асинхронный контекстный менеджер для работы с БД"""
    conn = await aiosqlite.connect('filters.db')
    conn.row_factory = aiosqlite.Row
    try:
        yield conn
        await conn.commit()
    except Exception as e:
        await conn.rollback()
        raise e
    finally:
        await conn.close()

async def get_user_filters(user_id: int) -> List[Dict]:
    """Асинхронное получение фильтров пользователя"""
    try:
        async with get_db_connection() as conn:
            cur = await conn.cursor()
            await cur.execute("SELECT * FROM filters WHERE user_id = ? ORDER BY expiry_date", (user_id,))
            rows = await cur.fetchall()
            return [dict(row) for row in rows]
    except Exception as e:
        logging.error(f"Ошибка при получении фильтров пользователя {user_id}: {e}")
        health_monitor.record_error()
        return []

async def get_filter_by_id(filter_id: int, user_id: int) -> Optional[Dict]:
    """Асинхронное получение фильтра по ID"""
    try:
        async with get_db_connection() as conn:
            cur = await conn.cursor()
            await cur.execute("SELECT * FROM filters WHERE id = ? AND user_id = ?", (filter_id, user_id))
            result = await cur.fetchone()
            return dict(result) if result else None
    except Exception as e:
        logging.error(f"Ошибка при получении фильтра {filter_id}: {e}")
        health_monitor.record_error()
        return None

async def get_all_users_stats() -> Dict:
    """Асинхронное получение статистики"""
    try:
        async with get_db_connection() as conn:
            cur = await conn.cursor()
            await cur.execute('''SELECT COUNT(DISTINCT user_id) as total_users, 
                                  COUNT(*) as total_filters,
                                  SUM(CASE WHEN expiry_date <= date('now') THEN 1 ELSE 0 END) as expired_filters,
                                  SUM(CASE WHEN expiry_date BETWEEN date('now') AND date('now', '+7 days') THEN 1 ELSE 0 END) as expiring_soon
                           FROM filters''')
            result = await cur.fetchone()
            return dict(result) if result else {'total_users': 0, 'total_filters': 0, 'expired_filters': 0, 'expiring_soon': 0}
    except Exception as e:
        logging.error(f"Ошибка при получении статистики: {e}")
        health_monitor.record_error()
        return {'total_users': 0, 'total_filters': 0, 'expired_filters': 0, 'expiring_soon': 0}

async def add_filter_to_db(user_id: int, filter_type: str, location: str, last_change: str, expiry_date: str, lifetime_days: int) -> bool:
    """Добавление фильтра в БД"""
    try:
        async with get_db_connection() as conn:
            cur = await conn.cursor()
            await cur.execute('''INSERT INTO filters 
                              (user_id, filter_type, location, last_change, expiry_date, lifetime_days) 
                              VALUES (?, ?, ?, ?, ?, ?)''',
                              (user_id, filter_type, location, last_change, expiry_date, lifetime_days))
            return True
    except Exception as e:
        logging.error(f"Ошибка при добавлении фильтра: {e}")
        health_monitor.record_error()
        return False

async def update_filter_in_db(filter_id: int, user_id: int, **kwargs) -> bool:
    """Обновление фильтра в БД"""
    try:
        if not kwargs:
            return False
        
        async with get_db_connection() as conn:
            cur = await conn.cursor()
            set_clause = ", ".join([f"{key} = ?" for key in kwargs.keys()])
            values = list(kwargs.values())
            values.extend([filter_id, user_id])
            
            await cur.execute(f"UPDATE filters SET {set_clause} WHERE id = ? AND user_id = ?", values)
            return cur.rowcount > 0
    except Exception as e:
        logging.error(f"Ошибка при обновлении фильтра {filter_id}: {e}")
        health_monitor.record_error()
        return False

async def delete_filter_from_db(filter_id: int, user_id: int) -> bool:
    """Удаление фильтра из БД"""
    try:
        async with get_db_connection() as conn:
            cur = await conn.cursor()
            await cur.execute("DELETE FROM filters WHERE id = ? AND user_id = ?", (filter_id, user_id))
            return cur.rowcount > 0
    except Exception as e:
        logging.error(f"Ошибка при удалении фильтра {filter_id}: {e}")
        health_monitor.record_error()
        return False

# ========== УЛУЧШЕНИЕ: УЛУЧШЕННАЯ ВАЛИДАЦИЯ ДАТ ==========
def try_auto_correct_date(date_str: str) -> Optional[datetime.date]:
    """Попытка автоматического исправления даты"""
    clean = re.sub(r'\D', '', date_str)
    
    if len(clean) == 6:  # ДДММГГ
        try:
            day, month, year = int(clean[:2]), int(clean[2:4]), int(clean[4:])
            if year < 100:
                year += 2000 if year < 50 else 1900
            return datetime(year, month, day).date()
        except ValueError:
            pass
    elif len(clean) == 8:  # ДДММГГГГ
        try:
            day, month, year = int(clean[:2]), int(clean[2:4]), int(clean[4:])
            return datetime(year, month, day).date()
        except ValueError:
            pass
    
    return None

def validate_date(date_str: str) -> datetime.date:
    """Улучшенная валидация даты с автокоррекцией"""
    date_str = date_str.strip()
    
    # Автозамена разделителей
    date_str = re.sub(r'[/\-]', '.', date_str)
    
    # Удаляем лишние символы, но оставляем точки
    date_str = re.sub(r'[^\d\.]', '', date_str)
    
    formats = ['%d.%m.%y', '%d.%m.%Y', '%d%m%y', '%d%m%Y', '%d.%m', '%d%m']
    
    for fmt in formats:
        try:
            if fmt in ['%d.%m', '%d%m']:
                # Добавляем текущий год
                date_obj = datetime.strptime(date_str, fmt).date()
                date_obj = date_obj.replace(year=datetime.now().year)
            elif fmt in ['%d%m%y', '%d%m%Y']:
                if len(date_str) in [6, 8]:
                    date_obj = datetime.strptime(date_str, fmt).date()
                else:
                    continue
            else:
                date_obj = datetime.strptime(date_str, fmt).date()
            
            today = datetime.now().date()
            max_past = today - timedelta(days=5*365)
            max_future = today + timedelta(days=1)
            
            if date_obj > max_future:
                raise ValueError("Дата не может быть в будущем")
            if date_obj < max_past:
                raise ValueError("Дата слишком старая (более 5 лет)")
                
            return date_obj
        except ValueError:
            continue
    
    # Попытка автоматического исправления
    corrected = try_auto_correct_date(date_str)
    if corrected:
        return corrected
    
    raise ValueError("Неверный формат даты. Используйте ДД.ММ.ГГ или ДД.ММ")

# ========== УЛУЧШЕНИЕ: ПРОВЕРКА ПРОСРОЧЕННЫХ ФИЛЬТРОВ ==========
async def check_expired_filters():
    """Проверка и уведомление о просроченных фильтрах"""
    try:
        async with get_db_connection() as conn:
            cur = await conn.cursor()
            await cur.execute("""
                SELECT DISTINCT user_id, filter_type, location, expiry_date 
                FROM filters 
                WHERE expiry_date <= date('now') 
                OR expiry_date BETWEEN date('now') AND date('now', '+7 days')
            """)
            expired_filters = await cur.fetchall()
            
            for filter_data in expired_filters:
                user_id = filter_data['user_id']
                expiry_date = datetime.strptime(str(filter_data['expiry_date']), '%Y-%m-%d').date()
                days_until = (expiry_date - datetime.now().date()).days
                
                if days_until <= 0:
                    message = (
                        f"🚨 <b>ВНИМАНИЕ! ФИЛЬТР ПРОСРОЧЕН</b>\n\n"
                        f"💧 <b>Тип:</b> {filter_data['filter_type']}\n"
                        f"📍 <b>Место:</b> {filter_data['location']}\n"
                        f"📅 <b>Срок истек:</b> {format_date_nice(expiry_date)}\n\n"
                        f"⚠️ <i>Немедленно замените фильтр!</i>"
                    )
                else:
                    message = (
                        f"⚠️ <b>ФИЛЬТР СКОРО ИСТЕЧЕТ</b>\n\n"
                        f"💧 <b>Тип:</b> {filter_data['filter_type']}\n"
                        f"📍 <b>Место:</b> {filter_data['location']}\n"
                        f"📅 <b>Срок истечения:</b> {format_date_nice(expiry_date)}\n"
                        f"⏰ <b>Осталось дней:</b> {days_until}\n\n"
                        f"💡 <i>Запланируйте замену заранее</i>"
                    )
                
                try:
                    await bot.send_message(user_id, message, parse_mode='HTML')
                    await asyncio.sleep(0.1)  # Защита от лимитов Telegram
                except Exception as e:
                    logging.error(f"Не удалось отправить уведомление пользователю {user_id}: {e}")
                    
    except Exception as e:
        logging.error(f"Ошибка при проверке просроченных фильтров: {e}")
        health_monitor.record_error()

# ========== УЛУЧШЕНИЕ: ЕЖЕДНЕВНЫЕ ОТЧЕТЫ ==========
async def send_daily_report():
    """Ежедневный отчет о состоянии бота"""
    try:
        health = await health_monitor.get_health_status()
        stats = await get_all_users_stats()
        
        report = (
            "📊 <b>ЕЖЕДНЕВНЫЙ ОТЧЕТ БОТА</b>\n\n"
            f"⏰ <b>Аптайм:</b> {health['uptime']}\n"
            f"📨 <b>Сообщений обработано:</b> {health['message_count']}\n"
            f"❌ <b>Ошибок:</b> {health['error_count']}\n"
            f"👥 <b>Активных пользователей:</b> {health['active_users']}\n"
            f"❤️ <b>Здоровье системы:</b> {health['health_score']:.1f}%\n\n"
            f"<b>СТАТИСТИКА БАЗЫ ДАННЫХ:</b>\n"
            f"👥 Пользователей: {stats['total_users']}\n"
            f"📦 Фильтров: {stats['total_filters']}\n"
            f"🔴 Просрочено: {stats['expired_filters']}\n"
            f"🟡 Скоро истечет: {stats['expiring_soon']}"
        )
        
        await bot.send_message(ADMIN_ID, report, parse_mode='HTML')
        logging.info("Ежедневный отчет отправлен администратору")
        
    except Exception as e:
        logging.error(f"Ошибка при отправке ежедневного отчета: {e}")

# ========== ОСТАЛЬНЫЕ НАСТРОЙКИ ==========
DEFAULT_LIFETIMES = {
    "магистральный sl10": 180,
    "магистральный sl20": 180,
    "гейзер": 365,
    "аквафор": 365
}

# УБРАН ЛИМИТ НА ФИЛЬТРЫ
MAX_FILTERS_PER_USER = 1000  # Очень высокий лимит, практически без ограничений

# States
class FilterStates(StatesGroup):
    waiting_filter_type = State()
    waiting_location = State()
    waiting_change_date = State()
    waiting_lifetime = State()

class MultipleFiltersStates(StatesGroup):
    waiting_filters_list = State()

class EditFilterStates(StatesGroup):
    waiting_filter_selection = State()
    waiting_field_selection = State()
    waiting_new_value = State()

class DeleteFilterStates(StatesGroup):
    waiting_filter_selection = State()

class ImportExportStates(StatesGroup):
    waiting_excel_file = State()

# ========== УЛУЧШЕНИЕ: АСИНХРОННАЯ ИНИЦИАЛИЗАЦИЯ БАЗЫ ==========
async def init_db():
    """Асинхронная инициализация базы данных"""
    try:
        async with get_db_connection() as conn:
            cur = await conn.cursor()
            
            # Проверяем существование таблицы
            await cur.execute("""
                SELECT name FROM sqlite_master 
                WHERE type='table' AND name='filters'
            """)
            table_exists = await cur.fetchone()
            
            if not table_exists:
                # Создаем таблицу
                await cur.execute('''
                    CREATE TABLE filters (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER,
                        filter_type TEXT,
                        location TEXT,
                        last_change DATE,
                        expiry_date DATE,
                        lifetime_days INTEGER,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                ''')
                
                # Создаем индексы
                await cur.execute('''CREATE INDEX idx_user_id ON filters(user_id)''')
                await cur.execute('''CREATE INDEX idx_expiry_date ON filters(expiry_date)''')
                logging.info("База данных успешно создана")
            else:
                logging.info("База данных уже существует")
                
    except Exception as e:
        logging.error(f"Критическая ошибка инициализации БД: {e}")
        # Создаем резервную копию при критической ошибке
        if os.path.exists('filters.db'):
            backup_name = f'filters_backup_critical_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
            try:
                shutil.copy2('filters.db', backup_name)
                logging.info(f"Создана критическая резервная копия: {backup_name}")
            except Exception as backup_error:
                logging.error(f"Не удалось создать резервную копию: {backup_error}")
        raise

# ========== ОСНОВНЫЕ ОБРАБОТЧИКИ ==========
@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    """Обработчик команды /start с rate limiting"""
    health_monitor.record_message(message.from_user.id)
    
    await message.answer(
        "🌟 <b>Фильтр-Трекер</b> 🤖\n\n"
        "💧 <i>Умный помощник для своевременной замены фильтров</i>\n\n"
        "📦 <b>Основные возможности:</b>\n"
        "• 📋 Просмотр всех ваших фильтров\n"
        "• ✨ Добавление новых фильтров\n"
        "• ⏳ Контроль сроков замены\n"
        "• ⚙️ Полное управление базой\n"
        "• 📊 Детальная статистика\n"
        "• 📤 Импорт/экспорт Excel\n"
        "• 🔔 Автоматические напоминания\n"
        "• 📊 Онлайн Excel с синхронизацией",
        reply_markup=get_main_keyboard(),
        parse_mode='HTML'
    )

@dp.message(Command("help"))
async def cmd_help(message: types.Message):
    """Обработчик команды /help"""
    help_text = (
        "🌟 <b>Фильтр-Трекер - Помощь</b> 🤖\n\n"
        "📋 <b>Основные команды:</b>\n"
        "• /start - Начать работу\n"
        "• /help - Показать справку\n"
        "• /status - Статус бота (админ)\n"
        "• /sync_excel - Синхронизировать Excel\n"
        "• /close_excel - Закрыть онлайн Excel\n"
        "• /excel_status - Статус Excel сессии\n\n"
        "💡 <b>Как использовать:</b>\n"
        "1. Добавьте фильтры через меню\n"
        "2. Следите за сроками замены\n"
        "3. Получайте уведомления\n"
        "4. Используйте онлайн Excel для удобства\n\n"
        "⚙️ <b>Управление:</b>\n"
        "• 📋 Мои фильтры - просмотр всех\n"
        "• ✨ Добавить фильтр - новый фильтр\n"
        "• ⚙️ Управление - редактирование\n"
        "• 📊 Статистика - ваша статистика\n"
        "• 📤 Импорт/Экспорт - работа с Excel\n\n"
        "❌ <b>Отмена операций:</b>\n"
        "Используйте кнопку '❌ Отмена' для отмены текущей операции"
    )
    await message.answer(help_text, parse_mode='HTML', reply_markup=get_main_keyboard())

@dp.message(F.text == "📋 Мои фильтры")
async def cmd_my_filters(message: types.Message):
    """Показать фильтры с rate limiting"""
    health_monitor.record_message(message.from_user.id)
    
    filters = await get_user_filters(message.from_user.id)
    
    if not filters:
        await message.answer(
            "📭 <b>У вас пока нет фильтров</b>\n\n"
            "💫 <i>Добавьте первый фильтр с помощью кнопки '✨ Добавить фильтр'</i>",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
        return
    
    today = datetime.now().date()
    response = "📋 <b>ВАШИ ФИЛЬТРЫ</b>\n\n"
    
    for f in filters:
        expiry_date = datetime.strptime(str(f['expiry_date']), '%Y-%m-%d').date()
        last_change = datetime.strptime(str(f['last_change']), '%Y-%m-%d').date()
        days_until_expiry = (expiry_date - today).days
        
        icon, status = get_status_icon_and_text(days_until_expiry)
        
        response += (
            f"{icon} <b>#{f['id']} {f['filter_type']}</b>\n"
            f"📍 {f['location']}\n"
            f"📅 Заменен: {format_date_nice(last_change)}\n"
            f"🗓️ Годен до: {format_date_nice(expiry_date)}\n"
            f"⏱️ Осталось дней: <b>{days_until_expiry}</b>\n"
            f"📊 Статус: <b>{status}</b>\n\n"
        )
    
    infographic = create_expiry_infographic(filters)
    await message.answer(response, parse_mode='HTML')
    await message.answer(infographic, parse_mode='HTML', reply_markup=get_main_keyboard())

@dp.message(F.text == "✨ Добавить фильтр")
async def cmd_add(message: types.Message):
    """Добавление фильтра"""
    await message.answer(
        "🔧 <b>Выберите тип добавления:</b>\n\n"
        "💡 <i>Можно добавить один фильтр или сразу несколько</i>",
        reply_markup=get_add_filter_keyboard(),
        parse_mode='HTML'
    )

@dp.message(F.text == "➕ Один фильтр")
async def cmd_add_single(message: types.Message, state: FSMContext):
    """Добавление одного фильтра"""
    await state.set_state(FilterStates.waiting_filter_type)
    await message.answer(
        "💧 <b>Выберите тип фильтра:</b>\n\n"
        "💡 <i>Используйте кнопки для быстрого выбора или введите свой вариант</i>",
        reply_markup=get_filter_type_keyboard(),
        parse_mode='HTML'
    )

@dp.message(FilterStates.waiting_filter_type)
async def process_filter_type(message: types.Message, state: FSMContext):
    """Обработка типа фильтра"""
    if message.text == "🔙 Назад":
        await state.clear()
        await message.answer("🔙 <b>Возврат в меню добавления</b>", reply_markup=get_add_filter_keyboard(), parse_mode='HTML')
        return
    
    filter_type = message.text.strip()
    
    # Если выбран "Другой тип фильтра", ждем ручной ввод
    if filter_type == "Другой тип фильтра":
        await message.answer(
            "💧 <b>Введите свой тип фильтра:</b>\n\n"
            "💡 <i>Например: Барьер, Атолл, Брита и т.д.</i>",
            reply_markup=get_back_keyboard(),
            parse_mode='HTML'
        )
        return
    
    # Валидация типа фильтра
    is_valid, error_msg = validate_filter_type(filter_type)
    if not is_valid:
        await message.answer(f"❌ {error_msg}\n\nПожалуйста, выберите тип фильтра еще раз:", reply_markup=get_filter_type_keyboard())
        return
    
    await state.update_data(filter_type=filter_type)
    await state.set_state(FilterStates.waiting_location)
    
    await message.answer(
        "📍 <b>Введите местоположение фильтра:</b>\n\n"
        "💡 <i>Например: Кухня, Ванная, Под раковиной и т.д.</i>",
        reply_markup=get_back_keyboard(),
        parse_mode='HTML'
    )

@dp.message(FilterStates.waiting_location)
async def process_location(message: types.Message, state: FSMContext):
    """Обработка местоположения"""
    if message.text == "🔙 Назад":
        await state.set_state(FilterStates.waiting_filter_type)
        await message.answer(
            "💧 <b>Выберите тип фильтра:</b>",
            reply_markup=get_filter_type_keyboard(),
            parse_mode='HTML'
        )
        return
    
    location = message.text.strip()
    
    # Валидация местоположения
    is_valid, error_msg = validate_location(location)
    if not is_valid:
        await message.answer(f"❌ {error_msg}\n\nПожалуйста, введите местоположение еще раз:", reply_markup=get_back_keyboard())
        return
    
    await state.update_data(location=location)
    await state.set_state(FilterStates.waiting_change_date)
    
    await message.answer(
        "📅 <b>Введите дату последней замены:</b>\n\n"
        "💡 <i>Формат: ДД.ММ.ГГ или ДД.ММ (текущий год)</i>\n"
        "📝 <i>Пример: 15.09.23 или 15.09</i>",
        reply_markup=get_back_keyboard(),
        parse_mode='HTML'
    )

@dp.message(FilterStates.waiting_change_date)
async def process_change_date(message: types.Message, state: FSMContext):
    """Обработка даты замены"""
    if message.text == "🔙 Назад":
        await state.set_state(FilterStates.waiting_location)
        await message.answer(
            "📍 <b>Введите местоположение фильтра:</b>",
            reply_markup=get_back_keyboard(),
            parse_mode='HTML'
        )
        return
    
    try:
        change_date = validate_date(message.text)
        await state.update_data(change_date=change_date)
        
        user_data = await state.get_data()
        filter_type = user_data.get('filter_type', '').lower()
        
        # Автоматическое определение срока службы
        lifetime = DEFAULT_LIFETIMES.get(filter_type, 180)
        
        await state.update_data(lifetime=lifetime)
        await state.set_state(FilterStates.waiting_lifetime)
        
        await message.answer(
            f"⏱️ <b>Срок службы фильтра:</b> {lifetime} дней\n\n"
            f"💡 <i>Автоматически определен для '{user_data['filter_type']}'</i>\n"
            f"📝 <i>Если хотите изменить, введите новое значение (в днях), или нажмите 'Пропустить'</i>",
            reply_markup=types.ReplyKeyboardMarkup(
                keyboard=[
                    [types.KeyboardButton(text="Пропустить")],
                    [types.KeyboardButton(text="🔙 Назад")]
                ],
                resize_keyboard=True
            ),
            parse_mode='HTML'
        )
        
    except ValueError as e:
        await message.answer(f"❌ {str(e)}\n\nПожалуйста, введите дату в правильном формате:", reply_markup=get_back_keyboard())

@dp.message(FilterStates.waiting_lifetime)
async def process_lifetime(message: types.Message, state: FSMContext):
    """Обработка срока службы"""
    if message.text == "🔙 Назад":
        await state.set_state(FilterStates.waiting_change_date)
        await message.answer(
            "📅 <b>Введите дату последней замены:</b>",
            reply_markup=get_back_keyboard(),
            parse_mode='HTML'
        )
        return
    
    user_data = await state.get_data()
    
    if message.text != "Пропустить":
        is_valid, error_msg, lifetime = validate_lifetime(message.text)
        if not is_valid:
            await message.answer(f"❌ {error_msg}\n\nПожалуйста, введите корректное число дней:")
            return
    else:
        lifetime = user_data['lifetime']
    
    # Расчет даты истечения
    change_date = user_data['change_date']
    expiry_date = change_date + timedelta(days=lifetime)
    
    # Сохранение в БД
    success = await add_filter_to_db(
        user_id=message.from_user.id,
        filter_type=user_data['filter_type'],
        location=user_data['location'],
        last_change=change_date.strftime('%Y-%m-%d'),
        expiry_date=expiry_date.strftime('%Y-%m-%d'),
        lifetime_days=lifetime
    )
    
    if success:
        await message.answer(
            f"✅ <b>ФИЛЬТР УСПЕШНО ДОБАВЛЕН!</b>\n\n"
            f"💧 <b>Тип:</b> {user_data['filter_type']}\n"
            f"📍 <b>Место:</b> {user_data['location']}\n"
            f"📅 <b>Замена:</b> {format_date_nice(change_date)}\n"
            f"🗓️ <b>Годен до:</b> {format_date_nice(expiry_date)}\n"
            f"⏱️ <b>Срок:</b> {lifetime} дней\n\n"
            f"💫 <i>Теперь я буду следить за сроком замены этого фильтра</i>",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
    else:
        await message.answer(
            "❌ <b>Ошибка при добавлении фильтра!</b>\n\n"
            "Пожалуйста, попробуйте еще раз.",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
    
    await state.clear()

# ========== ДОБАВЛЕНИЕ НЕСКОЛЬКИХ ФИЛЬТРОВ ==========
@dp.message(F.text == "📝 Несколько фильтров")
async def cmd_add_multiple(message: types.Message, state: FSMContext):
    """Добавление нескольких фильтров"""
    await state.set_state(MultipleFiltersStates.waiting_filters_list)
    await message.answer(
        "📝 <b>Введите несколько фильтров в формате:</b>\n\n"
        "💡 <i>Каждый фильтр с новой строки в формате:</i>\n"
        "<code>Тип фильтра; Местоположение; Дата замены</code>\n\n"
        "📝 <b>Пример:</b>\n"
        "<code>Магистральный SL10; Кухня; 15.09.23</code>\n"
        "<code>Гейзер; Ванная; 20.08.23</code>\n\n"
        "🚀 <b>Быстрые типы:</b> Магистральный SL10, Магистральный SL20, Гейзер, Аквафор\n\n"
        "💡 <i>Срок службы будет определен автоматически по типу фильтра</i>",
        reply_markup=get_back_keyboard(),
        parse_mode='HTML'
    )

@dp.message(MultipleFiltersStates.waiting_filters_list)
async def process_multiple_filters(message: types.Message, state: FSMContext):
    """Обработка нескольких фильтров"""
    if message.text == "🔙 Назад":
        await state.clear()
        await message.answer("🔙 <b>Возврат в меню добавления</b>", reply_markup=get_add_filter_keyboard(), parse_mode='HTML')
        return
    
    try:
        lines = message.text.strip().split('\n')
        added_count = 0
        errors = []
        
        for i, line in enumerate(lines, 1):
            parts = [part.strip() for part in line.split(';')]
            
            if len(parts) < 3:
                errors.append(f"Строка {i}: Недостаточно данных (нужно 3 части)")
                continue
            
            filter_type, location, date_str = parts[0], parts[1], parts[2]
            
            # Валидация типа фильтра
            is_valid_type, error_msg = validate_filter_type(filter_type)
            if not is_valid_type:
                errors.append(f"Строка {i}: {error_msg}")
                continue
            
            # Валидация местоположения
            is_valid_location, error_msg = validate_location(location)
            if not is_valid_location:
                errors.append(f"Строка {i}: {error_msg}")
                continue
            
            # Валидация даты
            try:
                change_date = validate_date(date_str)
            except ValueError as e:
                errors.append(f"Строка {i}: {str(e)}")
                continue
            
            # Определение срока службы
            lifetime = DEFAULT_LIFETIMES.get(filter_type.lower(), 180)
            expiry_date = change_date + timedelta(days=lifetime)
            
            # Сохранение в БД
            success = await add_filter_to_db(
                user_id=message.from_user.id,
                filter_type=filter_type,
                location=location,
                last_change=change_date.strftime('%Y-%m-%d'),
                expiry_date=expiry_date.strftime('%Y-%m-%d'),
                lifetime_days=lifetime
            )
            
            if success:
                added_count += 1
            else:
                errors.append(f"Строка {i}: Ошибка сохранения в БД")
        
        # Формирование ответа
        response = f"✅ <b>ДОБАВЛЕНО ФИЛЬТРОВ: {added_count}</b>\n\n"
        
        if errors:
            response += "❌ <b>Ошибки:</b>\n" + "\n".join(f"• {error}" for error in errors[:10])  # Показываем первые 10 ошибок
            if len(errors) > 10:
                response += f"\n\n... и еще {len(errors) - 10} ошибок"
        
        await message.answer(response, parse_mode='HTML', reply_markup=get_main_keyboard())
        
    except Exception as e:
        logging.error(f"Ошибка при обработке нескольких фильтров: {e}")
        await message.answer(
            "❌ <b>Произошла ошибка при обработке данных</b>\n\n"
            "Пожалуйста, проверьте формат и попробуйте еще раз.",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
    
    await state.clear()

@dp.message(F.text == "📊 Статистика")
async def cmd_stats(message: types.Message):
    """Показать статистику"""
    health_monitor.record_message(message.from_user.id)
    
    stats = await get_all_users_stats()
    user_filters = await get_user_filters(message.from_user.id)
    
    response = (
        "📊 <b>ВАША СТАТИСТИКА</b>\n\n"
        f"📦 <b>Всего фильтров:</b> {len(user_filters)}\n\n"
    )
    
    if is_admin(message.from_user.id):
        global_stats = (
            f"👑 <b>АДМИН СТАТИСТИКА</b>\n"
            f"👥 Пользователей: {stats['total_users']}\n"
            f"📦 Фильтров: {stats['total_filters']}\n"
            f"🔴 Просрочено: {stats['expired_filters']}\n"
            f"🟡 Скоро истечет: {stats['expiring_soon']}"
        )
        response += global_stats
    
    await message.answer(response, parse_mode='HTML', reply_markup=get_main_keyboard())

@dp.message(F.text == "⚙️ Управление фильтрами")
async def cmd_manage(message: types.Message):
    """Управление фильтрами"""
    health_monitor.record_message(message.from_user.id)
    
    filters = await get_user_filters(message.from_user.id)
    
    if not filters:
        await message.answer(
            "📭 <b>У вас пока нет фильтров для управления</b>\n\n"
            "💫 <i>Добавьте первый фильтр с помощью кнопки '✨ Добавить фильтр'</i>",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
        return
    
    await message.answer(
        "⚙️ <b>УПРАВЛЕНИЕ ФИЛЬТРАМИ</b>\n\n"
        "Выберите действие:",
        reply_markup=get_management_keyboard()
    )

# ========== РЕДАКТИРОВАНИЕ ФИЛЬТРОВ ==========
@dp.message(F.text == "✏️ Редактировать фильтр")
async def cmd_edit_filter(message: types.Message, state: FSMContext):
    """Начало редактирования фильтра"""
    filters = await get_user_filters(message.from_user.id)
    
    if not filters:
        await message.answer(
            "📭 <b>У вас пока нет фильтров для редактирования</b>",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
        return
    
    # Создаем клавиатуру с фильтрами
    builder = ReplyKeyboardBuilder()
    for f in filters:
        builder.button(text=f"#{f['id']} {f['filter_type']} - {f['location']}")
    builder.button(text="🔙 Назад")
    builder.adjust(1)
    
    await state.set_state(EditFilterStates.waiting_filter_selection)
    await message.answer(
        "📝 <b>Выберите фильтр для редактирования:</b>",
        reply_markup=builder.as_markup(resize_keyboard=True),
        parse_mode='HTML'
    )

@dp.message(EditFilterStates.waiting_filter_selection)
async def process_filter_selection_for_edit(message: types.Message, state: FSMContext):
    """Обработка выбора фильтра для редактирования"""
    if message.text == "🔙 Назад":
        await state.clear()
        await message.answer("🔙 <b>Возврат в меню управления</b>", reply_markup=get_management_keyboard(), parse_mode='HTML')
        return
    
    # Извлекаем ID фильтра из текста
    match = re.match(r'#(\d+)', message.text)
    if not match:
        await message.answer("❌ Неверный формат. Пожалуйста, выберите фильтр из списка.", reply_markup=get_back_keyboard())
        return
    
    filter_id = int(match.group(1))
    user_id = message.from_user.id
    
    # Проверяем существование фильтра
    filter_data = await get_filter_by_id(filter_id, user_id)
    if not filter_data:
        await message.answer("❌ Фильтр не найден.", reply_markup=get_back_keyboard())
        return
    
    # Сохраняем данные фильтра в состоянии
    await state.update_data(
        filter_id=filter_id,
        current_filter=filter_data
    )
    
    await state.set_state(EditFilterStates.waiting_field_selection)
    
    await message.answer(
        f"📝 <b>Редактирование фильтра #{filter_id}</b>\n\n"
        f"💧 <b>Тип:</b> {filter_data['filter_type']}\n"
        f"📍 <b>Место:</b> {filter_data['location']}\n"
        f"📅 <b>Замена:</b> {format_date_nice(datetime.strptime(str(filter_data['last_change']), '%Y-%m-%d'))}\n"
        f"⏱️ <b>Срок:</b> {filter_data['lifetime_days']} дней\n\n"
        "🔧 <b>Выберите поле для редактирования:</b>",
        reply_markup=get_edit_keyboard(),
        parse_mode='HTML'
    )

@dp.message(EditFilterStates.waiting_field_selection)
async def process_field_selection(message: types.Message, state: FSMContext):
    """Обработка выбора поля для редактирования"""
    if message.text == "🔙 Назад":
        await state.clear()
        await message.answer("🔙 <b>Возврат в меню управления</b>", reply_markup=get_management_keyboard(), parse_mode='HTML')
        return
    
    user_data = await state.get_data()
    filter_data = user_data['current_filter']
    
    field_map = {
        "💧 Тип фильтра": "filter_type",
        "📍 Местоположение": "location", 
        "📅 Дата замены": "last_change",
        "⏱️ Срок службы": "lifetime_days"
    }
    
    if message.text not in field_map:
        await message.answer("❌ Пожалуйста, выберите поле из списка.", reply_markup=get_edit_keyboard())
        return
    
    field = field_map[message.text]
    await state.update_data(editing_field=field)
    await state.set_state(EditFilterStates.waiting_new_value)
    
    current_value = filter_data[field]
    
    if field == "last_change":
        current_value = format_date_nice(datetime.strptime(str(current_value), '%Y-%m-%d'))
        prompt = "📅 <b>Введите новую дату замены:</b>\n\n💡 <i>Формат: ДД.ММ.ГГ или ДД.ММ</i>"
    elif field == "lifetime_days":
        prompt = f"⏱️ <b>Текущий срок службы:</b> {current_value} дней\n\nВведите новое значение (в днях):"
    else:
        prompt = f"✏️ <b>Текущее значение:</b> {current_value}\n\nВведите новое значение:"
    
    await message.answer(prompt, parse_mode='HTML', reply_markup=get_back_keyboard())

@dp.message(EditFilterStates.waiting_new_value)
async def process_new_value(message: types.Message, state: FSMContext):
    """Обработка нового значения для поля"""
    if message.text == "🔙 Назад":
        await state.set_state(EditFilterStates.waiting_field_selection)
        user_data = await state.get_data()
        filter_data = user_data['current_filter']
        
        await message.answer(
            f"📝 <b>Редактирование фильтра #{user_data['filter_id']}</b>\n\n"
            f"💧 <b>Тип:</b> {filter_data['filter_type']}\n"
            f"📍 <b>Место:</b> {filter_data['location']}\n"
            f"📅 <b>Замена:</b> {format_date_nice(datetime.strptime(str(filter_data['last_change']), '%Y-%m-%d'))}\n"
            f"⏱️ <b>Срок:</b> {filter_data['lifetime_days']} дней\n\n"
            "🔧 <b>Выберите поле для редактирования:</b>",
            reply_markup=get_edit_keyboard(),
            parse_mode='HTML'
        )
        return
    
    user_data = await state.get_data()
    filter_id = user_data['filter_id']
    field = user_data['editing_field']
    user_id = message.from_user.id
    
    try:
        if field == "filter_type":
            is_valid, error_msg = validate_filter_type(message.text)
            if not is_valid:
                await message.answer(f"❌ {error_msg}\n\nПожалуйста, введите корректное значение:", reply_markup=get_back_keyboard())
                return
            new_value = message.text
            
        elif field == "location":
            is_valid, error_msg = validate_location(message.text)
            if not is_valid:
                await message.answer(f"❌ {error_msg}\n\nПожалуйста, введите корректное значение:", reply_markup=get_back_keyboard())
                return
            new_value = message.text
            
        elif field == "last_change":
            try:
                change_date = validate_date(message.text)
                new_value = change_date.strftime('%Y-%m-%d')
                
                # Пересчитываем дату истечения
                filter_data = user_data['current_filter']
                expiry_date = change_date + timedelta(days=filter_data['lifetime_days'])
                await update_filter_in_db(filter_id, user_id, expiry_date=expiry_date.strftime('%Y-%m-%d'))
                
            except ValueError as e:
                await message.answer(f"❌ {str(e)}\n\nПожалуйста, введите дату в правильном формате:", reply_markup=get_back_keyboard())
                return
                
        elif field == "lifetime_days":
            is_valid, error_msg, lifetime = validate_lifetime(message.text)
            if not is_valid:
                await message.answer(f"❌ {error_msg}\n\nПожалуйста, введите корректное число дней:", reply_markup=get_back_keyboard())
                return
            new_value = lifetime
            
            # Пересчитываем дату истечения
            filter_data = user_data['current_filter']
            last_change = datetime.strptime(str(filter_data['last_change']), '%Y-%m-%d').date()
            expiry_date = last_change + timedelta(days=lifetime)
            await update_filter_in_db(filter_id, user_id, expiry_date=expiry_date.strftime('%Y-%m-%d'))
        
        # Обновляем поле в БД
        if field not in ["last_change", "lifetime_days"]:  # Эти поля уже обработаны выше
            success = await update_filter_in_db(filter_id, user_id, **{field: new_value})
        else:
            success = True
        
        if success:
            await message.answer(
                f"✅ <b>ФИЛЬТР ОБНОВЛЕН!</b>\n\n"
                f"🆔 <b>Фильтр #</b>{filter_id}\n"
                f"📝 <b>Поле обновлено:</b> {field}\n"
                f"💫 <b>Новое значение:</b> {new_value}",
                reply_markup=get_main_keyboard(),
                parse_mode='HTML'
            )
        else:
            await message.answer(
                "❌ <b>Ошибка при обновлении фильтра!</b>",
                reply_markup=get_main_keyboard(),
                parse_mode='HTML'
            )
            
    except Exception as e:
        logging.error(f"Ошибка при обновлении фильтра: {e}")
        await message.answer(
            "❌ <b>Произошла ошибка при обновлении</b>",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
    
    await state.clear()

# ========== УДАЛЕНИЕ ФИЛЬТРОВ ==========
@dp.message(F.text == "🗑️ Удалить фильтр")
async def cmd_delete_filter(message: types.Message, state: FSMContext):
    """Начало удаления фильтра"""
    filters = await get_user_filters(message.from_user.id)
    
    if not filters:
        await message.answer(
            "📭 <b>У вас пока нет фильтров для удаления</b>",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
        return
    
    # Создаем клавиатуру с фильтрами
    builder = ReplyKeyboardBuilder()
    for f in filters:
        builder.button(text=f"#{f['id']} {f['filter_type']} - {f['location']}")
    builder.button(text="🔙 Назад")
    builder.adjust(1)
    
    await state.set_state(DeleteFilterStates.waiting_filter_selection)
    await message.answer(
        "🗑️ <b>Выберите фильтр для удаления:</b>",
        reply_markup=builder.as_markup(resize_keyboard=True),
        parse_mode='HTML'
    )

@dp.message(DeleteFilterStates.waiting_filter_selection)
async def process_filter_deletion(message: types.Message, state: FSMContext):
    """Обработка удаления фильтра"""
    if message.text == "🔙 Назад":
        await state.clear()
        await message.answer("🔙 <b>Возврат в меню управления</b>", reply_markup=get_management_keyboard(), parse_mode='HTML')
        return
    
    # Извлекаем ID фильтра из текста
    match = re.match(r'#(\d+)', message.text)
    if not match:
        await message.answer("❌ Неверный формат. Пожалуйста, выберите фильтр из списка.", reply_markup=get_back_keyboard())
        return
    
    filter_id = int(match.group(1))
    user_id = message.from_user.id
    
    # Проверяем существование фильтра
    filter_data = await get_filter_by_id(filter_id, user_id)
    if not filter_data:
        await message.answer("❌ Фильтр не найден.", reply_markup=get_back_keyboard())
        return
    
    # Удаляем фильтр
    success = await delete_filter_from_db(filter_id, user_id)
    
    if success:
        await message.answer(
            f"✅ <b>ФИЛЬТР УДАЛЕН!</b>\n\n"
            f"🆔 <b>Фильтр #</b>{filter_id}\n"
            f"💧 <b>Тип:</b> {filter_data['filter_type']}\n"
            f"📍 <b>Место:</b> {filter_data['location']}\n"
            f"🗑️ <i>Фильтр успешно удален из базы данных</i>",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
    else:
        await message.answer(
            "❌ <b>Ошибка при удалении фильтра!</b>",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
    
    await state.clear()

# ========== ОБНОВЛЕННЫЙ ОНЛАЙН EXCEL С СИНХРОНИЗАЦИЕЙ ==========
@dp.message(F.text == "📊 Онлайн Excel")
async def cmd_online_excel(message: types.Message):
    """Улучшенный онлайн Excel с синхронизацией"""
    health_monitor.record_message(message.from_user.id)
    
    filters = await get_user_filters(message.from_user.id)
    
    if not filters:
        await message.answer(
            "📭 <b>У вас пока нет фильтров</b>\n\n"
            "💫 <i>Добавьте первый фильтр с помощью кнопки '✨ Добавить фильтр'</i>",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
        return
    
    try:
        # Создаем или обновляем сессию
        filepath = await excel_manager.create_session(message.from_user.id, filters)
        
        # Отправляем файл
        with open(filepath, 'rb') as file:
            await message.answer_document(
                types.BufferedInputFile(file.read(), filename="мои_фильтры_online.xlsx"),
                caption=(
                    "📊 <b>ОНЛАЙН EXCEL АКТИВИРОВАН</b>\n\n"
                    "💡 <b>Возможности онлайн режима:</b>\n"
                    "• 🔄 Автосинхронизация каждые 5 минут\n"
                    "• 📱 Редактирование в реальном времени\n"
                    "• 💾 Автосохранение изменений\n"
                    "• 🔔 Уведомления об обновлениях\n\n"
                    "🔄 <b>Команды управления:</b>\n"
                    "• /sync_excel - принудительная синхронизация\n"
                    "• /close_excel - закрыть онлайн сессию\n"
                    "• /excel_status - статус сессии"
                ),
                parse_mode='HTML',
                reply_markup=await get_online_excel_keyboard()
            )
        
    except Exception as e:
        logging.error(f"Ошибка создания онлайн Excel: {e}")
        await message.answer(
            "❌ <b>Ошибка при создании онлайн Excel!</b>\n\n"
            "Пожалуйста, попробуйте позже.",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )

# ========== КОМАНДЫ УПРАВЛЕНИЯ ONLINE EXCEL ==========
@dp.message(Command("sync_excel"))
async def cmd_sync_excel(message: types.Message):
    """Принудительная синхронизация онлайн Excel"""
    try:
        success = await excel_manager.sync_session(message.from_user.id)
        
        if success:
            await message.answer(
                "✅ <b>Excel файл синхронизирован!</b>\n\n"
                "💫 Все изменения из базы данных перенесены в ваш онлайн файл.",
                parse_mode='HTML',
                reply_markup=await get_online_excel_keyboard()
            )
        else:
            await message.answer(
                "❌ <b>Активная сессия не найдена!</b>\n\n"
                "Запустите онлайн Excel через меню управления.",
                parse_mode='HTML'
            )
            
    except Exception as e:
        logging.error(f"Ошибка синхронизации Excel: {e}")
        await message.answer(
            "❌ <b>Ошибка синхронизации!</b>",
            parse_mode='HTML'
        )

@dp.message(Command("close_excel"))
async def cmd_close_excel(message: types.Message):
    """Закрытие онлайн сессии Excel"""
    try:
        await excel_manager.close_session(message.from_user.id)
        await message.answer(
            "✅ <b>Онлайн сессия Excel закрыта</b>\n\n"
            "📁 Временный файл удален.\n"
            "💫 Вы можете создать новую сессию когда потребуется.",
            parse_mode='HTML',
            reply_markup=get_main_keyboard()
        )
    except Exception as e:
        logging.error(f"Ошибка закрытия Excel сессии: {e}")
        await message.answer(
            "❌ <b>Ошибка при закрытии сессии!</b>",
            parse_mode='HTML'
        )

@dp.message(Command("excel_status"))
async def cmd_excel_status(message: types.Message):
    """Статус онлайн Excel сессии"""
    try:
        session = excel_manager.active_sessions.get(message.from_user.id)
        
        if session:
            filters_count = len(session['filters'])
            last_update = session['last_update'].strftime('%d.%m.%Y %H:%M:%S')
            is_modified = "Да" if session['is_modified'] else "Нет"
            
            status_text = (
                "📊 <b>СТАТУС ONLINE EXCEL СЕССИИ</b>\n\n"
                f"📁 <b>Файл:</b> {session['filename']}\n"
                f"📦 <b>Фильтров в файле:</b> {filters_count}\n"
                f"🕒 <b>Последнее обновление:</b> {last_update}\n"
                f"✏️ <b>Изменения:</b> {is_modified}\n"
                f"🔄 <b>Автосинхронизация:</b> Каждые 5 минут\n\n"
                f"💡 <i>Сессия активна и синхронизируется автоматически</i>"
            )
        else:
            status_text = (
                "📊 <b>СТАТУС ONLINE EXCEL СЕССИИ</b>\n\n"
                "❌ <b>Активная сессия не найдена</b>\n\n"
                "💫 Запустите онлайн Excel через меню управления фильтрами."
            )
        
        await message.answer(status_text, parse_mode='HTML')
        
    except Exception as e:
        logging.error(f"Ошибка получения статуса Excel: {e}")
        await message.answer(
            "❌ <b>Ошибка получения статуса!</b>",
            parse_mode='HTML'
        )

# ========== INLINE ОБРАБОТЧИКИ ДЛЯ ONLINE EXCEL ==========
@dp.callback_query(F.data == "sync_excel")
async def callback_sync_excel(callback: types.CallbackQuery):
    """Обработчик синхронизации через inline кнопку"""
    try:
        success = await excel_manager.sync_session(callback.from_user.id)
        
        if success:
            await callback.message.edit_caption(
                caption=(
                    callback.message.caption + "\n\n✅ <b>Синхронизировано только что!</b>"
                ),
                parse_mode='HTML'
            )
            await callback.answer("Файл синхронизирован!")
        else:
            await callback.answer("Сессия не найдена!", show_alert=True)
            
    except Exception as e:
        logging.error(f"Ошибка inline синхронизации: {e}")
        await callback.answer("Ошибка синхронизации!", show_alert=True)

@dp.callback_query(F.data == "download_excel")
async def callback_download_excel(callback: types.CallbackQuery):
    """Обработчик скачивания актуальной версии"""
    try:
        filepath = await excel_manager.get_session_file(callback.from_user.id)
        
        if filepath and os.path.exists(filepath):
            with open(filepath, 'rb') as file:
                await callback.message.answer_document(
                    types.BufferedInputFile(file.read(), filename="актуальные_фильтры.xlsx"),
                    caption="📥 <b>Актуальная версия ваших фильтров</b>",
                    parse_mode='HTML'
                )
            await callback.answer("Файл отправлен!")
        else:
            await callback.answer("Файл не найден!", show_alert=True)
            
    except Exception as e:
        logging.error(f"Ошибка скачивания Excel: {e}")
        await callback.answer("Ошибка скачивания!", show_alert=True)

@dp.callback_query(F.data == "excel_stats")
async def callback_excel_stats(callback: types.CallbackQuery):
    """Статистика Excel файла"""
    try:
        session = excel_manager.active_sessions.get(callback.from_user.id)
        
        if session:
            filters = session['filters']
            today = datetime.now().date()
            
            expired = 0
            expiring_soon = 0
            normal = 0
            
            for f in filters:
                expiry_date = datetime.strptime(str(f['expiry_date']), '%Y-%m-%d').date()
                days_until = (expiry_date - today).days
                
                if days_until <= 0:
                    expired += 1
                elif days_until <= 7:
                    expiring_soon += 1
                else:
                    normal += 1
            
            stats_text = (
                "📈 <b>СТАТИСТИКА EXCEL ФАЙЛА</b>\n\n"
                f"📊 <b>Всего фильтров:</b> {len(filters)}\n"
                f"🟢 <b>Норма:</b> {normal}\n"
                f"🟡 <b>Скоро истечет:</b> {expiring_soon}\n"
                f"🔴 <b>Просрочено:</b> {expired}\n\n"
                f"💾 <b>Размер файла:</b> {os.path.getsize(session['filepath']) // 1024} КБ\n"
                f"🕒 <b>Обновлен:</b> {session['last_update'].strftime('%d.%m.%Y %H:%M')}"
            )
            
            await callback.message.answer(stats_text, parse_mode='HTML')
            await callback.answer()
        else:
            await callback.answer("Сессия не найдена!", show_alert=True)
            
    except Exception as e:
        logging.error(f"Ошибка статистики Excel: {e}")
        await callback.answer("Ошибка получения статистики!", show_alert=True)

@dp.callback_query(F.data == "close_excel")
async def callback_close_excel(callback: types.CallbackQuery):
    """Закрытие сессии через inline кнопку"""
    try:
        await excel_manager.close_session(callback.from_user.id)
        await callback.message.edit_caption(
            caption="❌ <b>Онлайн сессия Excel закрыта</b>\n\nФайл удален.",
            parse_mode='HTML'
        )
        await callback.answer("Сессия закрыта!")
        
    except Exception as e:
        logging.error(f"Ошибка закрытия Excel: {e}")
        await callback.answer("Ошибка закрытия сессии!", show_alert=True)

# ========== ИМПОРТ/ЭКСПОРТ ==========
@dp.message(F.text == "📤 Импорт/Экспорт")
async def cmd_import_export(message: types.Message):
    """Меню импорта/экспорта"""
    health_monitor.record_message(message.from_user.id)
    
    await message.answer(
        "📤 <b>ИМПОРТ/ЭКСПОРТ ДАННЫХ</b>\n\n"
        "💡 <b>Доступные функции:</b>\n"
        "• 📤 Экспорт в Excel - выгрузка всех ваших фильтров\n"
        "• 📥 Импорт из Excel - загрузка фильтров из файла\n"
        "• 📋 Шаблон Excel - скачать шаблон для заполнения\n\n"
        "⚠️ <i>Поддерживаются файлы .xlsx</i>",
        reply_markup=get_import_export_keyboard(),
        parse_mode='HTML'
    )

@dp.message(F.text == "📤 Экспорт в Excel")
async def cmd_export_excel(message: types.Message):
    """Экспорт фильтров в Excel"""
    filters = await get_user_filters(message.from_user.id)
    
    if not filters:
        await message.answer(
            "📭 <b>У вас пока нет фильтров для экспорта</b>",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
        return
    
    try:
        # Создаем DataFrame
        data = []
        today = datetime.now().date()
        
        for f in filters:
            expiry_date = datetime.strptime(str(f['expiry_date']), '%Y-%m-%d').date()
            last_change = datetime.strptime(str(f['last_change']), '%Y-%m-%d').date()
            days_until_expiry = (expiry_date - today).days
            
            icon, status = get_status_icon_and_text(days_until_expiry)
            
            data.append({
                'ID': f['id'],
                'Тип фильтра': f['filter_type'],
                'Местоположение': f['location'],
                'Дата последней замены': last_change.strftime('%d.%m.%Y'),
                'Дата истечения срока': expiry_date.strftime('%d.%m.%Y'),
                'Осталось дней': days_until_expiry,
                'Статус': status,
                'Срок службы (дни)': f['lifetime_days'],
                'Иконка статуса': icon
            })
        
        df = pd.DataFrame(data)
        
        # Создаем Excel файл в памяти
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Мои фильтры', index=False)
            
            # Настраиваем ширину колонок
            worksheet = writer.sheets['Мои фильтры']
            for idx, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).str.len().max(), len(col)) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 50)
        
        output.seek(0)
        
        # Отправляем файл
        await message.answer_document(
            types.BufferedInputFile(output.read(), filename=f"мои_фильтры_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"),
            caption=f"✅ <b>ЭКСПОРТ ЗАВЕРШЕН</b>\n\n📦 Экспортировано фильтров: {len(filters)}\n💫 Файл готов к использованию",
            parse_mode='HTML'
        )
        
    except Exception as e:
        logging.error(f"Ошибка при экспорте в Excel: {e}")
        await message.answer(
            "❌ <b>Ошибка при экспорте данных!</b>\n\n"
            "Пожалуйста, попробуйте позже.",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )

@dp.message(F.text == "📋 Шаблон Excel")
async def cmd_excel_template(message: types.Message):
    """Отправка шаблона Excel"""
    try:
        # Создаем шаблон DataFrame
        template_data = [
            {
                'Тип фильтра': 'Магистральный SL10',
                'Местоположение': 'Кухня',
                'Дата последней замены': '15.09.2023',
                'Срок службы (дни)': '180'
            },
            {
                'Тип фильтра': 'Гейзер',
                'Местоположение': 'Ванная',
                'Дата последней замены': '20.08.2023', 
                'Срок службы (дни)': '365'
            }
        ]
        
        df = pd.DataFrame(template_data)
        
        # Создаем Excel файл в памяти
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Шаблон', index=False)
            
            # Добавляем инструкции
            instructions = pd.DataFrame({
                'Инструкция': [
                    '1. Заполните данные в колонках A-D',
                    '2. Тип фильтра: Магистральный SL10, Магистральный SL20, Гейзер, Аквафор или другой',
                    '3. Местоположение: Кухня, Ванная, Под раковиной и т.д.',
                    '4. Дата замены в формате ДД.ММ.ГГГГ',
                    '5. Срок службы в днях (автоматически определится для популярных типов)',
                    '6. Сохраните файл и импортируйте в бота'
                ]
            })
            instructions.to_excel(writer, sheet_name='Инструкция', index=False)
            
            # Настраиваем ширину колонок
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for idx, col in enumerate(worksheet.iter_cols()):
                    max_len = 0
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_len:
                                max_len = len(str(cell.value))
                        except:
                            pass
                    worksheet.column_dimensions[chr(65 + idx)].width = min(max_len + 2, 50)
        
        output.seek(0)
        
        # Отправляем файл
        await message.answer_document(
            types.BufferedInputFile(output.read(), filename="шаблон_фильтров.xlsx"),
            caption="📋 <b>ШАБЛОН EXCEL</b>\n\n"
                   "💡 <i>Используйте этот шаблон для заполнения данных о фильтрах</i>\n\n"
                   "📝 <b>Как использовать:</b>\n"
                   "1. 📥 Скачайте шаблон\n"
                   "2. ✏️ Заполните данные\n"
                   "3. 📤 Загрузите файл через 'Импорт из Excel'",
            parse_mode='HTML'
        )
        
    except Exception as e:
        logging.error(f"Ошибка при создании шаблона Excel: {e}")
        await message.answer(
            "❌ <b>Ошибка при создании шаблона!</b>",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )

@dp.message(F.text == "📥 Импорт из Excel")
async def cmd_import_excel(message: types.Message, state: FSMContext):
    """Начало импорта из Excel"""
    await state.set_state(ImportExportStates.waiting_excel_file)
    await message.answer(
        "📥 <b>ИМПОРТ ИЗ EXCEL</b>\n\n"
        "💡 <b>Отправьте Excel файл (.xlsx) с данными о фильтрах</b>\n\n"
        "📋 <b>Формат файла:</b>\n"
        "• Колонка A: Тип фильтра\n"
        "• Колонка B: Местоположение\n"
        "• Колонка C: Дата замены (ДД.ММ.ГГГГ)\n"
        "• Колонка D: Срок службы (дни, опционально)\n\n"
        "⚠️ <i>Первая строка должна содержать заголовки</i>\n"
        "📎 <i>Или используйте готовый шаблон из меню</i>",
        reply_markup=get_back_keyboard(),
        parse_mode='HTML'
    )

@dp.message(ImportExportStates.waiting_excel_file, F.document)
async def process_excel_import(message: types.Message, state: FSMContext):
    """Обработка импорта Excel файла"""
    try:
        # Проверяем, что это Excel файл
        if not message.document.file_name.endswith(('.xlsx', '.xls')):
            await message.answer("❌ <b>Неверный формат файла!</b>\n\nОтправьте файл в формате .xlsx или .xls", parse_mode='HTML')
            return
        
        # Скачиваем файл
        file_info = await bot.get_file(message.document.file_id)
        downloaded_file = await bot.download_file(file_info.file_path)
        
        # Читаем Excel
        df = pd.read_excel(downloaded_file)
        
        # Проверяем необходимые колонки
        required_columns = ['Тип фильтра', 'Местоположение', 'Дата последней замены']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            await message.answer(
                f"❌ <b>В файле отсутствуют обязательные колонки:</b>\n{', '.join(missing_columns)}\n\n"
                f"💡 Используйте шаблон из меню 'Шаблон Excel'",
                parse_mode='HTML'
            )
            return
        
        # Обрабатываем данные
        added_count = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                filter_type = str(row['Тип фильтра']).strip()
                location = str(row['Местоположение']).strip()
                date_str = str(row['Дата последней замены']).strip()
                
                # Преобразуем дату из разных форматов
                if isinstance(date_str, str):
                    change_date = validate_date(date_str)
                else:
                    # Если это datetime объект
                    change_date = row['Дата последней замены'].date()
                
                # Определяем срок службы
                if 'Срок службы (дни)' in df.columns and pd.notna(row['Срок службы (дни)']):
                    lifetime = int(row['Срок службы (дни)'])
                else:
                    lifetime = DEFAULT_LIFETIMES.get(filter_type.lower(), 180)
                
                # Расчет даты истечения
                expiry_date = change_date + timedelta(days=lifetime)
                
                # Сохранение в БД
                success = await add_filter_to_db(
                    user_id=message.from_user.id,
                    filter_type=filter_type,
                    location=location,
                    last_change=change_date.strftime('%Y-%m-%d'),
                    expiry_date=expiry_date.strftime('%Y-%m-%d'),
                    lifetime_days=lifetime
                )
                
                if success:
                    added_count += 1
                else:
                    errors.append(f"Строка {idx+2}: Ошибка сохранения")
                    
            except Exception as e:
                errors.append(f"Строка {idx+2}: {str(e)}")
                continue
        
        # Формируем ответ
        response = f"✅ <b>ИМПОРТ ЗАВЕРШЕН</b>\n\n📦 Добавлено фильтров: {added_count}\n"
        
        if errors:
            response += f"❌ Ошибок: {len(errors)}\n"
            if len(errors) <= 5:
                response += "\n".join(f"• {error}" for error in errors)
            else:
                response += f"\nПоказаны первые 5 ошибок из {len(errors)}:\n" + "\n".join(f"• {error}" for error in errors[:5])
        
        await message.answer(response, parse_mode='HTML', reply_markup=get_main_keyboard())
        
    except Exception as e:
        logging.error(f"Ошибка при импорте Excel: {e}")
        await message.answer(
            "❌ <b>Ошибка при обработке файла!</b>\n\n"
            "Проверьте формат файла и попробуйте еще раз.",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
    
    await state.clear()

# ========== ОБРАБОТЧИКИ КНОПОК НАЗАД И ОТМЕНА ==========
@dp.message(F.text == "🔙 Назад")
async def cmd_back(message: types.Message, state: FSMContext):
    """Возврат в предыдущее меню"""
    current_state = await state.get_state()
    
    if current_state:
        await state.clear()
    
    await message.answer(
        "🔙 <b>Возврат в главное меню</b>",
        reply_markup=get_main_keyboard(),
        parse_mode='HTML'
    )

@dp.message(F.text == "❌ Отмена")
async def cmd_cancel(message: types.Message, state: FSMContext):
    """Отмена текущей операции"""
    current_state = await state.get_state()
    
    if current_state:
        await state.clear()
        await message.answer(
            "❌ <b>Операция отменена</b>",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )
    else:
        await message.answer(
            "ℹ️ <b>Нет активных операций для отмены</b>",
            reply_markup=get_main_keyboard(),
            parse_mode='HTML'
        )

# ========== КОМАНДЫ ДЛЯ АДМИНИСТРАТОРА ==========
@dp.message(Command("status"))
async def cmd_status(message: types.Message):
    """Команда для проверки статуса бота"""
    if not is_admin(message.from_user.id):
        await message.answer("❌ <b>Доступ запрещен!</b>", parse_mode='HTML')
        return
    
    health = await health_monitor.get_health_status()
    stats = await get_all_users_stats()
    
    status_report = (
        "🤖 <b>СТАТУС БОТА</b>\n\n"
        f"⏰ <b>Аптайм:</b> {health['uptime']}\n"
        f"📨 <b>Сообщений:</b> {health['message_count']}\n"
        f"❌ <b>Ошибок:</b> {health['error_count']}\n"
        f"👥 <b>Активных пользователей:</b> {health['active_users']}\n"
        f"❤️ <b>Здоровье системы:</b> {health['health_score']:.1f}%\n\n"
        f"<b>БАЗА ДАННЫХ:</b>\n"
        f"👥 Пользователей: {stats['total_users']}\n"
        f"📦 Фильтров: {stats['total_filters']}\n"
        f"🔴 Просрочено: {stats['expired_filters']}\n"
        f"🟡 Скоро истечет: {stats['expiring_soon']}"
    )
    
    await message.answer(status_report, parse_mode='HTML')

@dp.message(Command("reset_limits"))
async def cmd_reset_limits(message: types.Message):
    """Сброс rate limits (только для админа)"""
    if not is_admin(message.from_user.id):
        await message.answer("❌ <b>Доступ запрещен!</b>", parse_mode='HTML')
        return
    
    rate_limiter.user_requests.clear()
    await message.answer("✅ <b>Rate limits сброшены!</b>", parse_mode='HTML')

@dp.message(Command("backup"))
async def cmd_backup(message: types.Message):
    """Создание резервной копии (только для админа)"""
    if not is_admin(message.from_user.id):
        await message.answer("❌ <b>Доступ запрещен!</b>", parse_mode='HTML')
        return
    
    if backup_database():
        await message.answer("✅ <b>Резервная копия создана успешно!</b>", parse_mode='HTML')
    else:
        await message.answer("❌ <b>Ошибка при создании резервной копии!</b>", parse_mode='HTML')

# ========== УЛУЧШЕНИЕ: ОБНОВЛЕННЫЕ ФОНОВЫЕ ЗАДАЧИ ==========
async def schedule_daily_check():
    """Улучшенный планировщик ежедневных проверок"""
    while True:
        try:
            now = datetime.now()
            
            # Проверка просроченных фильтров каждый день в 10:00
            if now.hour == 10 and now.minute == 0:
                await check_expired_filters()
                
            # Резервное копирование в 3:00
            if now.hour == 3 and now.minute == 0:
                if backup_database():
                    logging.info("Резервная копия создана успешно")
                await asyncio.sleep(60)
            
            # Ежедневный отчет в 9:00
            if now.hour == 9 and now.minute == 0:
                await send_daily_report()
                await asyncio.sleep(60)
                
        except Exception as e:
            logging.error(f"Ошибка в фоновой задаче: {e}")
            health_monitor.record_error()
            await asyncio.sleep(300)
        
        await asyncio.sleep(60)

# ========== УЛУЧШЕНИЕ: ФОНОВЫЕ ЗАДАЧИ ДЛЯ ONLINE EXCEL ==========
async def schedule_online_excel_tasks():
    """Фоновые задачи для онлайн Excel"""
    # Запускаем авто-сохранение
    await excel_manager.start_auto_save()
    
    while True:
        try:
            # Синхронизация всех активных сессий каждые 5 минут
            await asyncio.sleep(300)  # 5 минут
            await excel_manager.sync_all_sessions()
            
            # Очистка старых сессий каждые 6 часов
            await excel_manager.cleanup_old_sessions()
            
        except Exception as e:
            logging.error(f"Ошибка в фоновых задачах Excel: {e}")
            await asyncio.sleep(60)

# ========== УЛУЧШЕНИЕ: ОБНОВЛЕННЫЙ ГЛОБАЛЬНЫЙ ОБРАБОТЧИК ОШИБОК ==========
@dp.errors()
async def errors_handler(update: types.Update, exception: Exception):
    """Улучшенный глобальный обработчик ошибок"""
    health_monitor.record_error()
    
    logging.error(f"Ошибка в update {update}: {exception}\n{traceback.format_exc()}")
    
    try:
        error_info = (
            f"❌ <b>ОШИБКА В БОТЕ</b>\n\n"
            f"🕒 <b>Время:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"🚨 <b>Тип:</b> {type(exception).__name__}\n"
            f"💬 <b>Ошибка:</b> {str(exception)[:1000]}\n\n"
            f"📊 <b>Статистика здоровья:</b>\n"
            f"• Сообщений: {health_monitor.message_count}\n"
            f"• Ошибок: {health_monitor.error_count}\n"
            f"• Здоровье: {(health_monitor.message_count - health_monitor.error_count) / max(1, health_monitor.message_count) * 100:.1f}%"
        )
        
        await bot.send_message(ADMIN_ID, error_info, parse_mode='HTML')
    except Exception as e:
        logging.error(f"Не удалось отправить сообщение об ошибке администратору: {e}")
    
    return True

# ========== ОБНОВЛЕННАЯ ФУНКЦИЯ ЗАПУСКА ==========
async def on_startup():
    """Улучшенные действия при запуске бота"""
    logging.info("Бот запущен")
    
    # Инициализация БД
    await init_db()
    
    # Создаем резервную копию при запуске
    if backup_database():
        logging.info("Резервная копия создана при запуске")
    
    # Запускаем фоновые задачи
    asyncio.create_task(schedule_daily_check())
    asyncio.create_task(schedule_online_excel_tasks())  # Новые задачи для Excel
    
    # Уведомляем администратора о запуске
    try:
        health = await health_monitor.get_health_status()
        await bot.send_message(
            ADMIN_ID, 
            f"🤖 <b>Бот успешно запущен!</b>\n\n"
            f"⏰ <b>Время запуска:</b> {health_monitor.start_time.strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"💫 <b>Статус:</b> Работает в нормальном режиме\n"
            f"📊 <b>Online Excel:</b> Активирован",
            parse_mode='HTML'
        )
    except Exception as e:
        logging.error(f"Не удалось отправить уведомление администратору: {e}")

# ========== ОБНОВЛЕННАЯ ФУНКЦИЯ ОСТАНОВКИ ==========
async def on_shutdown():
    """Действия при остановке бота"""
    logging.info("Бот останавливается")
    
    # Закрываем все активные сессии Excel
    for user_id in list(excel_manager.active_sessions.keys()):
        await excel_manager.close_session(user_id)
    
    try:
        await bot.send_message(
            ADMIN_ID,
            "🛑 <b>Бот остановлен</b>\n\n"
            f"⏰ <b>Время работы:</b> {datetime.now() - health_monitor.start_time}\n"
            f"📊 <b>Закрыто сессий Excel:</b> {len(excel_manager.active_sessions)}",
            parse_mode='HTML'
        )
    except Exception as e:
        logging.error(f"Не удалось отправить уведомление об остановке: {e}")

# Запуск бота
async def main():
    if not API_TOKEN:
        logging.error("Токен бота не найден! Установите переменную TELEGRAM_BOT_TOKEN")
        exit(1)
    
    await on_startup()
    try:
        await dp.start_polling(bot)
    except KeyboardInterrupt:
        logging.info("Бот остановлен пользователем")
        await on_shutdown()
    except Exception as e:
        logging.error(f"Критическая ошибка: {e}")
        await bot.send_message(ADMIN_ID, f"❌ Бот упал с ошибкой: {e}", parse_mode='HTML')
        raise

if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("Бот остановлен")
    except Exception as e:
        logging.critical(f"Критическая ошибка при запуске: {e}")
